# utils/export_excel.py
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension

def _apply_table_style(ws):
    # header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # apply header row style (first row)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # autofilter
    ws.auto_filter.ref = ws.dimensions

    # apply border to body
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            # wrap long text
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    # adjust column widths (basic)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
            except Exception:
                v = ""
            if len(v) > max_len:
                max_len = len(v)
        adjusted_width = min(60, max(10, (max_len + 2)))
        ws.column_dimensions[col_letter].width = adjusted_width


def df_to_excel_bytes(dfs):
    """
    Accepts either:
      - a single pandas DataFrame
      - OR a dictionary of {sheet_name: DataFrame}
    Returns Excel file bytes with formatting.
    """
    from pandas import DataFrame
    wb = Workbook()

    # Single DataFrame -> convert to dict
    if not isinstance(dfs, dict):
        dfs = {"Sheet1": dfs}

    first = True
    for sheet_name, df in dfs.items():
        if not isinstance(df, DataFrame):
            # try to convert pivot (which may be a DataFrame with index)
            try:
                df = df.reset_index()
            except Exception:
                df = DataFrame(df)

        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])

        # Write rows
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Protect against empty sheets
        if ws.max_row >= 1:
            _apply_table_style(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
